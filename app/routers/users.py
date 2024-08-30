"""
Руты для пользователей


CRUD с пользователями
Управление подпиской
"""
from operator import ne
import uuid, re, json
import openpyxl as xl 
from psycopg2.errors import UniqueViolation


from fastapi import (
    APIRouter, Depends, HTTPException, status, 
    Security, File, UploadFile
    )

from fastapi.encoders import jsonable_encoder

from typing import Annotated, Dict
from fastapi.responses import JSONResponse
from fastapi.security import OAuth2PasswordRequestForm, SecurityScopes
from datetime import datetime
from sqlalchemy import desc, asc, not_, or_
from app import Tags

from app.models import (
    Users, Orders, Session, engine, Roles, Permissions, 
    Address, UsersAddress, BoxTypes, OrderStatusHistory,
    OrderStatuses, ROLE_CUSTOMER_NAME, ROLE_TELEGRAM_BOT_NAME,
    Regions, UserRefreshTokens
    )

from app.auth import (
    oauth2_scheme, pwd_context, get_password_hash, 
    verify_password, create_access_token,
    get_current_active_user, get_current_user,
    ACCESS_TOKEN_EXPIRE_MINUTES, SECRET_KEY, ALGORITHM,
    create_refresh_token, get_current_user_refresh
)

from app.validators import (
    UserSignUp as UserSignUpSchema,
    UserLogin as UserLoginSchema,
    UserCreationValidator as UserCreationData,
    UserUpdateValidator as UserUpdateData,
    UserOut, OrderOut
)

from app.utils import is_valid_uuid, get_lang_long_from_text_addres
from app import logger

from passlib.context import CryptContext
from datetime import timedelta
from jose import jwt
from dotenv import load_dotenv

router = APIRouter()


@router.get('/users', tags=[Tags.admins, Tags.managers], responses={
    200: {
        "description": "Получение всех пользователей",
        "content": {
                "application/json": {
                    "example": 
                        [   
                            {
                                "full_name": 'null',
                                "date_created": "2023-11-29T00:57:57.654800",
                                "phone_number": 'null',
                                "email": 'null',
                                "telegram_id": 7643079034697,
                                "id": 1
                            }
                        ]
                }
        }
    }
})
async def get_all_users(
        current_user: Annotated[UserLoginSchema, Security(get_current_user, scopes=["manager"])],
        role_name: str = None,
        only_bot_users: bool = False,

        #Применимо для клиентов (На самом деле для всех)
        with_orders: bool = False,
        with_active_orders: bool = False,
        with_inactive_orders: bool = False, #все кроме активных
        has_orders: bool = False,


        #Применимо для всех
        limit: int = 5,
        page: int = 0,
        show_deleted: bool = True
    ):
    """
    Получение пользователей по фильтрам:

    - **role_name**: Фильтр по наличию роли у пользователя 
    - **only_bot_users**: Возвращать ли только пользователей, привязавших/начавших пользоваться ботом
    - **with_orders**: Возвращать информацию о пользователях вместе с их заявками
    - **with_active_orders**: bool - возвращать пользователей только с активными заявками
    - **limit**: кол-во пользователей на запрос
    - **page**: номер страницы
    - **show_deleted**: показывать удалённых пользвателей
    - **has_orders**: показывать только пользователей, у которых есть хотя бы одна заявка
    """

    with Session(engine, expire_on_commit=False) as session:
        query = session.query(Users)

        roles_query = session.query(Roles).all()
        roles_list = [role.role_name for role in roles_query]

        if not show_deleted:
            query = query.filter(Users.deleted_at == None)

        if only_bot_users:
            query = query.filter(or_(Users.link_code == None, not_(Users.telegram_username == None)))

        if (role_name != None) and (role_name in roles_list):
            roles_user_query = session.query(Users.id).\
                join(Permissions, Permissions.user_id == Users.id).\
                join(Roles, Roles.id == Permissions.role_id).\
                where(Roles.role_name == role_name).subquery()

            query = query.filter(Users.id.in_(roles_user_query))

        if has_orders:
            only_users_with_orders_query = session.query(Users.id).\
                join(Orders, Orders.from_user == Users.id).subquery()
            query = query.filter(Users.id.in_(only_users_with_orders_query))

        if with_active_orders:
            only_users_with_active_orders_query = session.query(Users.id).\
                join(Orders, Orders.from_user == Users.id).filter(or_(
                Orders.status == OrderStatuses.status_accepted_by_courier().id,
                Orders.status == OrderStatuses.status_default().id,
                Orders.status == OrderStatuses.status_processing().id,
                Orders.status == OrderStatuses.status_awating_confirmation().id,
                Orders.status == OrderStatuses.status_confirmed().id,
                Orders.status == OrderStatuses.status_awaiting_payment().id,
                Orders.status == OrderStatuses.status_payed().id
            )).subquery()
            query = query.filter(Users.id.in_(only_users_with_active_orders_query))
        

        global_user_count = query.count()
        if limit == 0:
            users = query.order_by(asc(Users.date_created)).all()
        else:
            users = query.order_by(asc(Users.date_created)).offset(page  * limit).limit(limit).all()

        total = len(users)
        data = []

        for user in users:
            scopes_query = session.query(Permissions, Roles.role_name).filter_by(user_id=user.id).join(Roles).all()
            user.roles = [role.role_name for role in scopes_query]
            user_parent_data = jsonable_encoder(user.__dict__)
            user_data = UserOut(**user_parent_data)
         
            if with_orders:
                orders = session.query(Orders, Address, BoxTypes, OrderStatuses).\
                    join(Address, Address.id == Orders.address_id).\
                    outerjoin(BoxTypes, BoxTypes.id == Orders.box_type_id).\
                    join(OrderStatuses, OrderStatuses.id == Orders.status).\
                    where(Orders.from_user == user.id).order_by(asc(Orders.date_created))

                if with_active_orders:
                    orders = orders.filter(or_(
                        Orders.status == OrderStatuses.status_accepted_by_courier().id,
                        Orders.status == OrderStatuses.status_default().id,
                        Orders.status == OrderStatuses.status_processing().id,
                        Orders.status == OrderStatuses.status_awating_confirmation().id,
                        Orders.status == OrderStatuses.status_confirmed().id,
                        Orders.status == OrderStatuses.status_awaiting_payment().id,
                        Orders.status == OrderStatuses.status_payed().id
                    ))

                orders = orders.all()                    

                user_data.orders = []
                for order in orders:
                    order_data = order[0].__dict__
                    del order_data['manager_info']
                    order_data = OrderOut(**order[0].__dict__)

                    try:
                        order_data.address_data = order[1]
                    except IndexError: 
                        order_data.address_data = None

                    try:
                        order_data.box_data = order[2]
                    except IndexError:
                        order_data.box_data = None

                    try:
                        order_data.status_data = order[3]
                    except IndexError:
                        order_data.status_data = None

                    user_data.orders.append(order_data)

            data.append(user_data)

        return {
            "count": total,
            "global_count": global_user_count,
            "data": data
        }


@router.post('/users', tags=[Tags.managers, Tags.admins])
async def create_user(
    current_user: Annotated[UserLoginSchema, Security(get_current_user, scopes=["manager"])],
    new_user_data: UserCreationData,
    send_email: bool = True
)->UserOut:
    """
    Ручное создание нового пользователя
    - **send_email**: Отправить ли пользвателю письмо с кодом связи
    """

    with Session(engine, expire_on_commit=False) as session:
        query_user = session.query(Users).filter_by(email=new_user_data.email).first()
        if query_user: 
            return JSONResponse({
                "detail": "Email already taken",
            }, status_code=400)

        if not (new_user_data.telegram_id == None):
            query_user = session.query(Users).\
                    filter_by(telegram_id=new_user_data.telegram_id).first()

            if query_user:
                return JSONResponse({
                    "detail": "Telegram id is already taken",
                }, status_code=400)


        password_plain = ''
        print(new_user_data.password)
        if new_user_data.password == None:
            password_plain = str(uuid.uuid4())[:10] 
        else:
            password_plain = str(new_user_data.password)

        print(password_plain)
        new_user_data.password = get_password_hash(password_plain)
        
        new_user_data = new_user_data.model_dump()
        user_role = new_user_data["role"]
        del new_user_data["role"]

        if send_email:
            pass

        new_user = Users(**new_user_data)

        #Фикс: при flush uuid остаётся в сесси и не перегенерируется, т.е получаем Exception на unique field'е 
        new_user.link_code = str(uuid.uuid4())[:10] 

        session.add(new_user)
        session.flush()
        session.refresh(new_user)

        #Если админ - добавить все роли?

        for role in user_role:
            role_query = Roles.get_role(role)
            if role_query:

                user_role = Permissions(
                    user_id = new_user.id,
                    role_id = Roles.get_role(role).id
                )

                session.add(user_role)

        session.commit()

        scopes_query = session.query(Permissions, Roles.role_name).\
                filter_by(user_id=new_user.id).join(Roles).all()

        scopes = [role.role_name for role in scopes_query]

        #TODO: Отправка приглашения на почту
        user_parent_data = jsonable_encoder(new_user.__dict__)
        user_out_data = UserOut(**user_parent_data)
        user_out_data.roles = scopes
        user_out_data.password_plain = password_plain

        return user_out_data


@router.delete('/users', tags=[Tags.admins, Tags.managers])
async def delete_user(
    current_user: Annotated[UserLoginSchema, Security(get_current_user, scopes=["manager"])],
    tg_id: int | None = None,
    user_id: uuid.UUID | None = None 
):
    """
    Удаление пользователя (Отправляется в disabled)
    """
    if (not tg_id) and (not user_id):
        return JSONResponse({
            "message": "Required at least one type id in arguments"
        }, status_code=422)

    with Session(engine, expire_on_commit=False) as session:
        user_query = None
        if tg_id:
            user_query = session.query(Users).filter_by(telegram_id = tg_id).where(deleted_at = None).first()
        elif user_id:
            user_query = session.query(Users).filter_by(id = user_id).where(Users.deleted_at == None).first()

        if not user_query:
            return JSONResponse({
                "message": "No such user"
                }, status_code=404)

        user_query.deleted_at = datetime.now()
        session.add(user_query)
        session.commit()

        return


@router.put('/user', tags=[Tags.managers, Tags.admins])
async def update_user_data(
    current_user: Annotated[UserLoginSchema, Security(get_current_user, scopes=["manager"])],
    new_user_data: UserUpdateData
)->UserOut:
    """
    Обновление данных пользователя админом или менеджером
    """
    with Session(engine, expire_on_commit=False) as session:
        user_query = Users.get_user(new_user_data.user_id)
        flag_email_same = False
        flag_phone_same = False

        if not user_query:
            return JSONResponse({
                "message": "No such user"
                }, status_code=404)

        for attr, value in new_user_data.model_dump(exclude_unset=True).items():
            # print(attr)
            if attr == 'email':
                if user_query.email == value:
                    flag_email_same = True

            if attr == 'phone_number':
                if user_query.phone_number == value:
                    flag_phone_same = True

            if attr == 'password' and value:
                setattr(user_query, attr, get_password_hash(value))
                continue
            
            if attr == 'roles' and value:
                delete_query = session.query(Permissions).filter(Permissions.user_id == user_query.id).delete()
                for role in new_user_data.roles:
                    role_query = Roles.get_role(role)
                    if role_query:
                        user_role = Permissions(
                            user_id = user_query.id,
                            role_id = Roles.get_role(role).id
                        )

                        session.add(user_role)

                continue

            setattr(user_query, attr, value)

        session.add(user_query)
        try:
            session.commit()
        except Exception as err:
            logger.debug(err)
            if 'users_email_key' in str(err):
                return JSONResponse({
                        "detail": 'Почта уже занята', 
                        "code": 422
                    },
                    status_code=422
                )
            elif 'phone' in str(err):
                return JSONResponse({
                        "detail": 'Телефон уже занят', 
                        "code": 423
                    },
                    status_code=423
                )

        user_parent_data = jsonable_encoder(user_query.__dict__)
        user_data = UserOut(**user_parent_data)
        scopes_query = session.query(Permissions, Roles.role_name).\
                filter_by(user_id=user_query.id).join(Roles).all()
        user_data.roles = [role.role_name for role in scopes_query]
        
        if flag_email_same:
            user_data.code = 205

        if flag_phone_same:
            user_data.code = 204

        return user_data


@router.post('/users/reset_password')
async def reset_user_password(
    user_id: uuid.UUID,
    current_user: Annotated[UserLoginSchema, Security(get_current_user, scopes=["admin"])],
):
    import secrets
    import string
    import os 

    alphabet = string.ascii_letters + string.digits
    password = ''.join(secrets.choice(alphabet) for i in range(20))  

    with Session(engine, expire_on_commit=False) as session:
        user = Users.get_user(user_id)
        if not user:
            raise HTTPException(
                detail='Пользователь не найден',
                status_code=404
            )

        setattr(user, "password", get_password_hash(password))
        session.add(user)
        session.commit()

    from fastapi_mail import FastMail, MessageSchema, ConnectionConfig
    from pydantic import EmailStr

    conf = ConnectionConfig(
        MAIL_USERNAME=os.getenv("MAIL_USERNAME"),
        MAIL_PASSWORD=os.getenv("MAIL_PASSWORD"),
        MAIL_FROM=os.getenv("MAIL_FROM"),
        MAIL_PORT=os.getenv("MAIL_PORT"),
        MAIL_SERVER=os.getenv("MAIL_SERVER"),
        MAIL_STARTTLS=os.getenv("MAIL_STARTTLS"),
        MAIL_SSL_TLS=os.getenv("MAIL_SSL_TLS"),
        USE_CREDENTIALS=os.getenv("USE_CREDENTIALS"),
    )

    async def send_email(emails: list[EmailStr], subject: str, message: str):
        message = MessageSchema(
            subject=subject, recipients=emails, body=message, subtype="html"
        )
        fm = FastMail(conf)
        await fm.send_message(message)

    await send_email(
        emails=[user.email],
        subject='Смена пароля',
        message=f"Ваш пароль был изменён. Новый пароль: {password}"
    )

    #fastapi_mail=1.4.1

    return JSONResponse({
        "new_password": password
    })


@router.get('/users/bot_linked', tags=[Tags.bot, Tags.users])
async def get_bot_users():
    """
    Получение пользователей тг бота, или пользователей сайта связавших профиль с тг
    """
    with Session(engine, expire_on_commit=False) as session:
        bot_users = session.query(Users).where(Users.link_code == None, Users.telegram_id is not None).all()
        return bot_users


@router.post('/users/signup', tags=[Tags.users])
async def signup(signup_data: UserSignUpSchema):
    """
    Регистрация клиента
    """

    with Session(engine, expire_on_commit=False) as session:
        query = session.query(Users).filter_by(email=signup_data.username).first()
        if query:
            return JSONResponse({
                "message": "Email already taken"
            }, status_code=409)

        new_user = Users(
            email=signup_data.username,
            password=get_password_hash(signup_data.password)
        )  

        session.add(new_user)
        session.flush()
        session.refresh(new_user)

        new_user.link_code = str(uuid.uuid4())[:10] 
        role_query = Roles.get_role(ROLE_CUSTOMER_NAME)

        if role_query:
            user_role = Permissions(
                user_id = new_user.id,
                role_id = role_query.id
            )

            session.add(user_role)
        else:
            raise HTTPException(status_code=503, detail="Internal server error")

        session.commit()

        return new_user


@router.post('/token', tags=[Tags.users])
async def login(login_data: Annotated[OAuth2PasswordRequestForm, Depends()]):

    with Session(engine, expire_on_commit=False) as session:
        query = session.query(Users).filter_by(email=login_data.username).first()
        if not query: 
            return JSONResponse({
                "message": "Invalid password or email"
            }, status_code=400)

        if not verify_password(login_data.password, query.password):
            return JSONResponse({
                "message": "Invalid password or email"
            }, 400)

        #TODO: Проверка на активность

        scopes_query = session.query(Permissions, Roles.role_name).filter_by(user_id=query.id).join(Roles).all()

        scopes = [role.role_name for role in scopes_query]

        expires = timedelta(minutes=int(ACCESS_TOKEN_EXPIRE_MINUTES))
        refresh_expires = timedelta(minutes=int(ACCESS_TOKEN_EXPIRE_MINUTES)*20)
        if ROLE_TELEGRAM_BOT_NAME in scopes:
            expires = None

        token = create_access_token(
            data={
                "sub": login_data.username,
                "internal_id": str(query.id),
                "scopes": scopes
            }, 
            expires_delta=expires)

        refresh_token = create_refresh_token(
            data={
                "internal_id": str(query.id),
                "scopes": scopes
            }, expires_delta=refresh_expires
        )
        new_refresh_token = UserRefreshTokens(
            user_id=query.id,
            token=refresh_token
        )
        session.add(new_refresh_token)
        session.commit()

        return JSONResponse({
            "access_token": token,
            "refresh_token": refresh_token,
            "token_type": "bearer"
        })


@router.post('/token/refresh', tags=[Tags.users])
async def refresh_access_token(
    current_user: Annotated[UserLoginSchema, Security(get_current_user_refresh)]
):
    """
    получить новый токен доступа, если текущий не истёк
    """
    with Session(engine, expire_on_commit=False) as session:

        query = session.query(Users).filter_by(id=current_user.id).first()
        scopes_query = session.query(Permissions, Roles.role_name).filter_by(user_id=query.id).join(Roles).all()

        scopes = [role.role_name for role in scopes_query]

        expires = timedelta(minutes=int(ACCESS_TOKEN_EXPIRE_MINUTES))
        if ROLE_TELEGRAM_BOT_NAME in scopes:
            expires = None
        
        found = False
        for token in current_user.refresh_tokens:
            if current_user.refresh_token == token:
                found = True
                delete_query = session.query(UserRefreshTokens).filter_by(token=token).delete()
                session.commit()

        if not found:
            return JSONResponse({
                "message": "not authenticated"
            }, status_code=401)

        token = create_access_token(
            data={
                "sub": query.email,
                "internal_id": str(query.id),
                "scopes": scopes
            }, 
            expires_delta=expires)

        refresh_expires = timedelta(minutes=int(ACCESS_TOKEN_EXPIRE_MINUTES)*20)
        refresh_token = create_refresh_token(
            data={
                "internal_id": str(query.id),
                "scopes": scopes
            }, expires_delta=refresh_expires
        )
        new_refresh_token = UserRefreshTokens(
            user_id=query.id,
            token=refresh_token
        )
        session.add(new_refresh_token)
        session.commit()

        return JSONResponse({
            "access_token": token,
            "refresh_token": refresh_token,
            "token_type": "bearer"
        })


@router.get("/users/me", tags=[Tags.users], responses={
    200: {
        "description": "Получение информации об авторизованном пользователе (себе)",
        "content": {
            "application/json": {
                "example": {
  "user_data": {
    "email": "user3@example.com",
    "id": "43f96b7c-c417-4be1-9be8-857bf9df8acb",
    "telegram_id": 55455,
    "telegram_username": 'null',
    "phone_number": 'null',
    "firstname": 'null',
    "secondname": 'null',
    "orders": 'null',
    "roles": 'null',
    "deleted_at": 'null',
    "link_code": "e08d1fb4-6"
  },
  "token_data": {
    "sub": "user3@example.com",
    "internal_id": "43f96b7c-c417-4be1-9be8-857bf9df8acb",
    "scopes": [
      "customer",
      "admin",
      "manager",
      "courier",
      "bot"
    ]
  }
                }
            }
        }
    },
    401: {
        "description": "Не удалось авторизировать пользователя",
        "content": {
                "application/json": {
                    "detail": "Not authenticated"
                }
            },
    }
})
async def get_current_user_info(
    current_user: Annotated[UserLoginSchema, Security(get_current_user)]
):
    """
    Получить информацию о текущем пользователе по токену
    """
    token_data = jwt.decode(token=current_user.access_token, key=SECRET_KEY, algorithms=ALGORITHM)

    with Session(engine, expire_on_commit=False) as session:
        query = session.query(Users).filter_by(email=current_user.username).first()

        scopes_query = session.query(Permissions, Roles.role_name).filter_by(user_id=query.id).join(Roles).all()
        scopes = [role.role_name for role in scopes_query]

        query.roles = scopes
        user_parent_data = jsonable_encoder(query.__dict__)
        user_data = UserOut(**user_parent_data)
        return {
            "user_data": user_data,
            "token_data": token_data
            }


@router.get('/user/{user_id}/info', tags=[Tags.users, Tags.admins])
async def get_user_data(
    user_id: uuid.UUID,
    current_user: Annotated[UserLoginSchema, Security(get_current_user, scopes=["admin"])],
    with_orders: bool = False
)->UserOut:
    """
    Получить данные пользователя админом
    """
    with Session(engine, expire_on_commit=False) as session:
        user_query = session.query(Users).where(Users.id == user_id).first()
        if not user_query:
            return JSONResponse({
                "message": "Not found"
            }, status_code=404)

        scopes_query = session.query(Permissions, Roles.role_name).filter_by(user_id=user_query.id).join(Roles).all()
        scopes = [role.role_name for role in scopes_query]

        user_query.roles = scopes
        user_parent_data = jsonable_encoder(user_query.__dict__)
        return_data = UserOut(**user_parent_data)

        if with_orders:
            orders = session.query(Orders).\
                    enable_eagerloads(False).\
                    where(Orders.from_user == user_query.id).\
                    order_by(asc(Orders.date_created)).all()

            orders_out = []
            for order in orders:
                parent_data = jsonable_encoder(order)
                order_data = OrderOut(**parent_data)

                order_data.tg_id = user_query.telegram_id
                order_data.address_data = order.address
                order_data.box_data = order.box
                order_data.status_data = session.query(OrderStatuses).filter_by(id=order.status).first()

                orders_out.append(order_data)

            return_data.orders = orders_out

        return return_data


@router.get('/user/me', tags=[Tags.bot, Tags.users])
async def get_user_info(
    tg_id: int,
    bot: Annotated[UserLoginSchema, Security(get_current_user, scopes=["bot"])]
)->UserOut:

    with Session(engine, expire_on_commit=False) as session:
        query = session.query(Users).filter_by(telegram_id=tg_id).first()
        if not query:
            return JSONResponse({
                "message": "Not found"
            }, status_code=404)

        scopes_query = session.query(Permissions, Roles.role_name).filter_by(user_id=query.id).join(Roles).all()
        scopes = [role.role_name for role in scopes_query]

        user_parent_data = jsonable_encoder(query.__dict__)
        return_data = UserOut(**user_parent_data)
        return_data.roles = scopes

        return return_data


@router.get('/users/check', tags=[Tags.admins, Tags.users])
async def check_user_roles(
    current_user: Annotated[UserLoginSchema, Security(get_current_user, scopes=["admin"])]
):
    """
    Проверить базу данных пользователей на пользователей без роли и добавить им роль customer
    """
    with Session(engine, expire_on_commit=False) as session:
        user_qeury = session.query(Users).all()
        for user in user_qeury:
            scopes_query = session.query(Permissions, Roles.role_name).filter_by(user_id=user.id).join(Roles).all()
            scopes = [role.role_name for role in scopes_query]
            if len(scopes) < 1:
                print(f"User {user.id} has no roles")
                print(scopes)
                if ROLE_CUSTOMER_NAME not in scopes:
                    user_role = Permissions(
                        user_id = user.id,
                        role_id = Roles.customer_role()
                    )
                    session.add(user_role)

        session.commit()


@router.post('/users/import/file', tags=[Tags.admins])
async def import_clients(
    file: UploadFile,
    current_user: Annotated[UserLoginSchema, Security(get_current_user, scopes=["admin"])]
    ):
    """
    Импорт клиентов и заявки с историей из excel таблицы
    """
    with Session(engine, expire_on_commit=False) as session:
    
        wb_obj = xl.load_workbook(file.file)
        sheet_obj = wb_obj.active

        users = []
        none_count = 0 

        #импорт самих клиентов, адресов, тарифов, подписки, интервалов

        # for row in range(1, sheet_obj.max_row+1):
            # print(f"row number {row} ")

        added_count = 0
        error_data = []

        for row in range(1, 160):
            #3840 24
            #print(sheet_obj.max_row, sheet_obj.max_column)
            cell_obj = sheet_obj.cell(row, 1)

            #Считаем кол-во пустых ячеек подряд
            if cell_obj.value == None:
                none_count +=1
            else:
                none_count = 0

            if not re.match(r'[\d](.*)', str(cell_obj.value)):
                #Если слишком много пустых ячеек - завершаем обработку, т.к.
                #файл может иметь тысячи пустых строк
                if none_count > 30:                
                    break
                continue

            # print(get_lang_long_from_text_addres(sheet_obj.cell(row,7).value))
            try:
                phone_number = int(sheet_obj.cell(row, 11).value)
            except Exception as err:

                error_data.append({
                    "message": "No phone number detected",
                    "row": row
                })

                continue
            
            #TODO: Обновление данных пользователя, если он уже существовал?
            new_user = session.query(Users).filter_by(phone_number = str(phone_number)).first()
            if not new_user:
                new_user = Users(
                    phone_number = int(sheet_obj.cell(row, 11).value),
                    firstname = sheet_obj.cell(row, 9).value,
                    link_code = str(uuid.uuid4())[:8]
                )

                session.add(new_user)
                session.flush()
                session.refresh(new_user)
            else:
                error_data.append({
                    "message": f"Пользователь с номером {phone_number} уже существует",
                    "row": row
                })
                continue
            
            new_address = session.query(Address).filter_by(address = sheet_obj.cell(row,7).value).\
                where(Address.deleted_at == None).\
                join(UsersAddress, UsersAddress.address_id == Address.id).\
                where(UsersAddress.user_id == new_user.id).\
                first()

            if not new_address:
                longitude, latitude = await get_lang_long_from_text_addres(sheet_obj.cell(row,7).value)
                if (latitude == None) or (longitude == None):
                    error_data.append({
                        "message": f"No data found for address {sheet_obj.cell(row, 7).value}",
                        "row": row
                    })
                    continue
                
                region = Regions.get_by_coords(latitude, longitude)
                if not region:
                    region = Regions.get_by_name(sheet_obj.cell(row, 4).value)
                    if not region:
                        error_data.append({
                            "message": f"No region found for {sheet_obj.cell(row, 4).value}",
                            "row": row
                        })
                        continue

                new_address = Address(
                    region_id = region.id,
                    distance_from_mkad = sheet_obj.cell(row, 6).value,
                    point_on_map = sheet_obj.cell(row, 8).value,
                    address = sheet_obj.cell(row, 7).value,
                    detail = sheet_obj.cell(row, 10).value,
                    main = True,
                    longitude = longitude,
                    latitude = latitude
                )

                session.add(new_address)
                session.flush()
                session.refresh(new_address)

            user_role = Permissions(
                user_id = new_user.id,
                role_id = Roles.customer_role()
            )

            count = session.query(Orders.id).where(Orders.from_user == new_user.id).count()
            new_order = Orders(
                from_user   = new_user.id,
                address_id  = new_address.id,
                box_type_id = BoxTypes.test_type().id,
                box_count   = 1,
                status      = OrderStatuses.status_default().id,
                date_created = datetime.now(),
                user_order_num = count + 1
            )

            session.add(new_order)
            session.add(user_role)
            session.flush()
            session.refresh(new_order)

            status_update = OrderStatusHistory(
                order_id = new_order.id,
                status_id = OrderStatuses.status_default().id,
                date = datetime.now()
            )

            user_address = UsersAddress(
                user_id=new_user.id,
                address_id=new_address.id,
            )

            session.add(user_address)
            session.add(status_update)
            session.add(new_order)

            try:
                session.commit()
            finally:
                added_count+=1

    return JSONResponse({
        "added_count": added_count,
        "error_count": len(error_data),
        "errors_detail": error_data
    })
        

@router.post('/users/phone-numbers/fix')
async def fix_phone_numbers(
    current_user: Annotated[UserLoginSchema, Security(get_current_user, scopes=["admin"])]
):
    from pydantic_extra_types.phone_numbers import PhoneNumber
    from pydantic import  BaseModel
    from typing import Optional

    class PhoneCheckValidator(BaseModel):
        phone_number: Optional[PhoneNumber] = None

    with Session(engine, expire_on_commit=False) as session:
        user_query = session.query(Users).all()

        for user in user_query:
            if user.phone_number == None:
                continue
            

            old_phone_number = str(user.phone_number)
            if '.0' in old_phone_number:
                old_phone_number = old_phone_number.replace('.0', '')

            print(f'old number: {old_phone_number}')
            new_phone_number = None

            try:
                if old_phone_number[0] == '8':
                    old_phone_number = "+7"+str(old_phone_number)[1::]
                    new_phone_number = PhoneCheckValidator(phone_number=old_phone_number)
                else: 
                    new_phone_number = PhoneCheckValidator(phone_number=old_phone_number)
            except Exception as err:
                user.phone_number = None

            if new_phone_number:
                check_existance = session.query(Users).filter_by(phone_number=new_phone_number.phone_number).first()
                if check_existance:
                    print("Phone exists")
                    user.phone_number = None
                else:
                    user.phone_number = new_phone_number.phone_number
            else:
                user.phone_number = None


            print(f"new: {new_phone_number}")
            print("")

            session.add(user)

        session.commit()